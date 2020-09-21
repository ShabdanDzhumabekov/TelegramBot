import psycopg2.errors as ps_err
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime



class DowloandError(Exception):
    def __init__(self, message):
        self.message = message


def save_marks_to_book(database):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Баллы студентов'
    ws['A4'] = 'Номер зачетки'
    ws['B4'] = 'Фамилия'
    ws['B3'] = 'ДЗ №'
    ws['B1'] = 'Под номером дз - дата дз'
    database.cursor.execute('''
    Select record_num, surname, name
    FROM sort''')
    students = database.cursor.fetchall()
    i = 5
    for line in students:
        j = 1
        ws.cell(i, j, line[0])
        j += 1
        ws.cell(i, j, line[1] + ' ' + line[2][0] + '.')
        i += 1
    database.cursor.execute('''
    select task_id, date_given from tasks''')
    num_of_tasks = database.cursor.fetchall()
    j = 3
    for num in num_of_tasks:
        marks = database.find_all_marks_with(num[0])
        i = 3
        ws.cell(i, j, num[0])  # номер дз
        i += 1
        ws.cell(i, j, num[1])  # дата дз
        i += 1
        dictionary_marks = {}
        for iter in marks:
            dictionary_marks.update({iter[0]: iter[1]})
        for stud in students:
            ws.cell(i, j, dictionary_marks[stud[0]])
            i += 1
        j += 1
    wb.save('marks.xlsx')
    wb.close()


def download_marks_from_book(database, file):
    wb = load_workbook(filename=file)
    ws = wb.active
    curr_cell = (4, 1)  # A4
    if 'Номер зачетки' != ws.cell(4, 1).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (4, 2)
    if 'Фамилия' != ws.cell(4, 2).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (3, 1)
    if 'ДЗ №' != ws.cell(3, 2).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    database.cursor.execute('''
        Select record_num, surname, name
        FROM sort''')
    students = database.cursor.fetchall()
    formatted_student = []
    for stud in students:
        formatted_student.append((stud[0], str(stud[1]) + ' ' + str(stud[2][0]) + '.'))
    for rows in ws.iter_rows(min_row=5, max_col=2, values_only=True):
        if rows not in formatted_student:
            raise DowloandError('Ошибка для {}. Отсутсвует в списке студентов'.format(rows))
    database.cursor.execute('''select task_id, date_given from tasks''')
    tasks = database.cursor.fetchall()
    for col in ws.iter_cols(min_row=3, min_col=3, max_row=4, values_only=True):
        if isinstance(col[1], datetime.datetime) and isinstance(col[0], int):
            formatted_col = (col[0], datetime.date(col[1].year, col[1].month, col[1].day))
        else:
            raise DowloandError('Ошибка для ДЗ {}. Неверный формат'.format(col))
        if formatted_col not in tasks:
            raise DowloandError('Ошибка для ДЗ №{} от {}. Отсутсвует в списке ДЗ'
                                .format(formatted_col[0], str(formatted_col[1])))

    def add_mark_of_task(database, task_id, record_num, mark):
        database.cursor.execute('''
        UPDATE done_tasks
        SET mark = %s
        WHERE task_id = %s AND record_num = %s''', (mark, task_id, record_num))

    try:
        j = 3  # номер дз
        for col in ws.iter_cols(min_row=5, min_col=3, values_only=True):
            task_id = ws.cell(3, j).value
            i = 5  # зачетка
            for mark in col:
                if isinstance(mark, int):
                    add_mark_of_task(database, task_id, ws.cell(i, 1).value, mark)
                else:
                    raise DowloandError('Ошибка в ячейке ({},{}). Введено не число'
                                        .format(i, j))
                i += 1
            j += 1
        database.connection.commit()
    except ps_err.CheckViolation:
        raise DowloandError('Ошибка в ячейке ({},{}). Балл должен быть > 0'
                            .format(i, j))
    wb.close()


def from_excel_to_database_of_students_list(database, file):
    wb = load_workbook(filename=file)
    ws = wb.active
    # Проверка excel файла на соответствующий формат
    curr_cell = (1, 1)  # A1
    if 'Номер зачетки' != ws.cell(1, 1).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 2)  # B1
    if 'Фамилия' != ws.cell(1, 2).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 3)  # C1
    if 'Имя' != ws.cell(1, 3).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 4)  # D1
    if 'Отчество' != ws.cell(1, 4).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    # Запрос в БД для взятия номеров зачеток
    database.cursor.execute('''
        select record_num
        FROM students''')
    rns = database.cursor.fetchall()
    # f = False  # флажок на проверку существования студента в БД
    try:
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            f = False
            for rn in rns:
                # Проверка на существования студента в БД (проверяются номера зачеток)
                if row[0] == rn[0]:
                    f = True
                    break
            # Если студент не существует в БД, заносим его туда
            if not f:
                database.cursor.execute('''
                    insert into students(name, surname, midname, record_num)
                    values (%s, %s, %s, %s)''', (row[2], row[1], row[3], row[0]))
                #database.add_student_out_id(row[2], row[1], row[3], row[0])
            else:
                database.cursor.execute('''
                    update students
                    set name = %s, surname = %s, midname = %s
                    where record_num = %s''', (row[2], row[1], row[3], row[0]))
                #database.update_student_with_record(row[0], *(row[2], row[1], row[3]))
        database.connection.commit()
    except ps_err.UniqueViolation:
        raise DowloandError('Существуют повторяющиеся номера зачеток!')
    except ps_err.NotNullViolation:
        raise DowloandError('Недопустимо оставлять фамилию или имя пустым!')
    wb.close()


def from_database_to_excel_of_students_list(database):
    wb = Workbook()
    ws = wb.active
    # Создаем первую строку имен столбцов
    ws.title = 'Список студентов'
    ws['A1'] = 'Номер зачетки'
    ws['B1'] = 'Фамилия'
    ws['C1'] = 'Имя'
    ws['D1'] = 'Отчество'
    # Запрос на БД для взятия данных
    database.cursor.execute('''
	Select record_num, surname, name, midname
	FROM students''')
    students = database.cursor.fetchall()
    i = 2  # Начинаем сохранение данных со второй строки
    for row in students:
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
        ws.cell(i, 3, row[2])
        ws.cell(i, 4, row[3])
        i += 1
    wb.save('students.xlsx')
    wb.close()


def delete_students_from_database(database, xl_file):
    wb = load_workbook(xl_file)
    ws = wb.active
    # Проверка excel файла на соответствующий формат
    curr_cell = (1, 1)  # A1
    if 'Номер зачетки' != ws.cell(1, 1).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 2)  # B1
    if 'Фамилия' != ws.cell(1, 2).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 3)  # C1
    if 'Имя' != ws.cell(1, 3).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    curr_cell = (1, 4)  # D1
    if 'Отчество' != ws.cell(1, 4).value:
        raise DowloandError('Ошибка в ячейке {}. Несоответствие формату'.format(curr_cell))
    # Удаляем студентов из бд
    deleted_list = [] #список user_id удаленных
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        stud = database.find_student_with_record(row[0])
        database.delete_student_with_record(row[0])
        if stud:
            if stud[4]:
                deleted_list.append(stud[4])
        #database.cursor.execute(''' DELETE FROM students WHERE record_num=%s ''' % row[0])
        #database.connection.commit()
    return deleted_list
